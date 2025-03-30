import openpyxl 
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



customer_sheet = openpyxl.load_workbook('dados_clientes.xlsx')
customer_page = customer_sheet['Sheet1']

closed_sheet = openpyxl.load_workbook('planilha fechamento.xlsx')
closed_page = closed_sheet['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

def format_text_split (text, index):
    return text.split()[index]

def get_element_text(driver, xpath):
    """Função para buscar o texto de um elemento pelo XPath."""
    sleep(1)  # Pequena pausa para garantir que o elemento carregue
    return driver.find_element(By.XPATH, xpath).text

for line in customer_page.iter_rows(min_row=2, values_only=True):
    name, price, cpf, expiration_date = line    
    # Inserindo CPF e realizando a consulta
    field_search = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "cpfInput")))
    field_search.clear()
    sleep(1)
    field_search.send_keys(cpf)
    sleep(1)
    button_consult = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    button_consult.click()
    sleep(5)
    
    status = get_element_text(driver, "//span[@id='statusLabel']")
    
    if status == 'em dia':
        payment_date = get_element_text(driver, "//p[@id='paymentDate']")
        method_payment = get_element_text(driver, "//p[@id='paymentMethod']")
        
        payment_date = format_text_split(payment_date, 3)
        method_payment = format_text_split(method_payment, 3)
        
        print([name, price, cpf, expiration_date, status, payment_date, method_payment])
        closed_page.append([name, price, cpf, expiration_date, status, payment_date, method_payment])
        closed_sheet.save('planilha fechamento.xlsx')
    else:
        print([name, price, cpf, expiration_date, 'Pendente'])

        closed_page.append([name, price, cpf, expiration_date, 'Pendente'])
        closed_sheet.save('planilha fechamento.xlsx')